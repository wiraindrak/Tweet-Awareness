from __future__ import division
from nltk.metrics import ConfusionMatrix

__author__ = 'WiraIndra'

import os
import pickle
import re
import xlrd
import openpyxl
import math

class DictionaryHandler:
    @staticmethod
    def loadDict_words():
        path_dict = os.path.dirname(__file__)
        dir_dict = os.path.join(path_dict, 'dictionary')
        dict_words = open(os.path.join(dir_dict, 'dict_vocab.txt'), 'r').read().splitlines()
        return dict_words

    @staticmethod
    def loadDict_noise():
        path_dict = os.path.dirname(__file__)
        dir_dict = os.path.join(path_dict, 'dictionary')
        dict_stopwords = open(os.path.join(dir_dict, 'dict_noise.txt'), 'r').read().splitlines()
        return dict_stopwords

    @staticmethod
    def loadPickle():
        path_dict = os.path.dirname(__file__)
        dir_binary = os.path.join(path_dict, 'binary')
        pickle_awareness = open(os.path.join(dir_binary, 'pickle_awareness.pickle'), 'rb')
        classifier_awareness = pickle.load(pickle_awareness)
        pickle_awareness.close()
        return classifier_awareness

class Preprocessing:
    @staticmethod
    def remove_username(tweet):
        return re.sub(r'@\w+', '', tweet)

    @staticmethod
    def remove_hashtag(tweet):
        return re.sub(r'#\w+', '', tweet)

    @staticmethod
    def remove_url(tweet):
        def regex_or(*items):
            r = '|'.join(items)
            r = '(' + r + ')'
            return r

        def pos_lookahead(r):
            return '(?=' + r + ')'

        def neg_lookahead(r):
            return '(?!' + r + ')'

        def optional(r):
            return '(%s)?' % r

        PunctChars = r'''['".?!,:;]'''
        Punct = '%s+' % PunctChars
        Entity = '&(amp|lt|gt|quot);'

        UrlStart1 = regex_or(r'https?://?', r'www\.')
        CommonTLDs = regex_or('com', 'co\\.uk', 'org', 'net', 'info', 'ca')
        UrlStart2 = r'[a-z0-9\.-]+?' + r'\.' + CommonTLDs + pos_lookahead(r'[/ \W\b]')
        UrlBody = r'[^ \t\r\n<>]*?'
        UrlExtraCrapBeforeEnd = '%s+?' % regex_or(PunctChars, Entity)
        UrlEnd = regex_or(r'\.\.+', r'[<>]', r'\s', '$')
        Url = (r'\b' +
               regex_or(UrlStart1, UrlStart2) +
               UrlBody +
               pos_lookahead(optional(UrlExtraCrapBeforeEnd) + UrlEnd))

        Url_RE = re.compile("(%s)" % Url, re.U | re.I)
        tweet = re.sub(Url_RE, "", tweet)

        return tweet

    @staticmethod
    def stemming(word):
        # affix di-
        word = re.sub('^di', '', word)

        # # affix -x
        # word = re.sub(r'([a-z0-9]+)x$', r'\1', word)

        # affix -ny
        word = re.sub(r'([a-z0-9]+)ny$', r'\1', word)

        # affix -nya
        word = re.sub(r'([a-z0-9]+)nya$', r'\1', word)

        # affix -nk
        word = re.sub(r'([a-z0-9]+)nk$', r'\1ng', word)

        # affix -dh
        word = re.sub(r'([a-z0-9]+)dh$', r'\1t', word)

        # # affix -ku
        # word = re.sub(r'([a-z0-9]+)ku$', r'\1', word)
        #
        # # affix -mu
        # word = re.sub(r'([a-z0-9]+)mu$', r'\1', word)

        # twin word
        word = re.sub(r'([a-z0-9]+)2$', r'\1', word)
        word = re.sub(r'([a-z0-9]+)2([a-z0-9]*)', r'\1', word)
        word = re.sub(r'([a-z0-9]+)-([a-z0-9]*)', r'\1', word)

        # substring oe
        word = re.sub('oe', 'u', word)

        # remove number
        word = re.sub(r'[0-9]', '', word)

        return word

    @staticmethod
    def convert_numberToLetter(word):
        if(re.match(r'[a-zA-Z]+[0-9]+[0-9a-zA-Z]*', word)):
            word = re.sub('12', 'r', word)
            word = re.sub('13', 'b', word)
            word = re.sub('0', 'o', word)
            word = re.sub('1', 'i', word)
            word = re.sub('2', 'r', word)
            word = re.sub('3', 'e', word)
            word = re.sub('4', 'a', word)
            word = re.sub('5', 's', word)
            word = re.sub('6', 'g', word)
            word = re.sub('7', 'j', word)
            word = re.sub('8', 'b', word)
            word = re.sub('9', 'g', word)

            return word

        elif(re.match(r'[0-9]+[a-zA-Z]+[0-9a-zA-Z]*', word)):
            word = re.sub('12', 'r', word)
            word = re.sub('13', 'b', word)
            word = re.sub('0', 'o', word)
            word = re.sub('1', 'i', word)
            word = re.sub('2', 'r', word)
            word = re.sub('3', 'e', word)
            word = re.sub('4', 'a', word)
            word = re.sub('5', 's', word)
            word = re.sub('6', 'g', word)
            word = re.sub('7', 'j', word)
            word = re.sub('8', 'b', word)
            word = re.sub('9', 'g', word)
            return word
        else:
            return word

    @staticmethod
    def convert_symbolToLetter(word):
        if(re.match(r'[a-zA-Z0-9]+[\@\!\$]+[\@\!\$0-9a-zA-Z]*', word)):
            word = re.sub('\@', 'a', word)
            word = re.sub('\!', 'i', word)
            word = re.sub('\$', 's', word)

            return word
        elif(re.match(r'[\@\!\$]+[a-zA-Z0-9]+[\@\!\$0-9a-zA-Z]*', word)):
            word = re.sub('\!', 'i', word)
            word = re.sub('\$', 's', word)

            return word
        else:
            return word

    @staticmethod
    def remove_multipleLetter(word):
        word = re.sub(r'([a-z])\1{2,}', r'\1', word)
        word = re.sub(r'([a-z][a-z])\1{2,}', r'\1', word)

        # Exceptions: A = maaf, E = seenaknya, G = nggak, I = diimbangi, K = diletakkan, L = Innalillahi, N = kecepatannya,  O = kooperatif, P = happy, R = sparring, S = massal, T = settingan
        word = re.sub(r'([cdfhjmquvwxyz])\1', r'\1', word)
        return word

    @staticmethod
    def remove_sign(tweet):
        tweet = tweet.replace('.', '')
        tweet = tweet.replace(',', '')
        tweet = tweet.replace('"', '')
        tweet = tweet.replace("'", "")
        tweet = tweet.replace('!', '')
        tweet = tweet.replace('/', '')
        tweet = tweet.replace('(', '')
        tweet = tweet.replace(')', '')
        tweet = tweet.replace(':', '')
        tweet = tweet.replace(';', '')
        tweet = tweet.replace('*', '')
        tweet = tweet.replace('[', '')
        tweet = tweet.replace(']', '')
        tweet = tweet.replace('|', '')
        tweet = tweet.replace('&', '')
        tweet = tweet.replace('%', '')
        tweet = tweet.replace('#', '')
        tweet = tweet.replace('$', '')
        tweet = tweet.replace('^', '')
        tweet = tweet.replace('+', '')
        tweet = tweet.replace(' - ', ' ')
        tweet = tweet.replace('- ', ' ')
        tweet = tweet.replace(' -', ' ')
        tweet = tweet.replace('_', '')
        tweet = tweet.replace('=', '')
        tweet = tweet.replace('{', '')
        tweet = tweet.replace('}', '')
        tweet = tweet.replace('~', '')
        tweet = tweet.replace('@', '')
        tweet = tweet.replace('<', '')
        tweet = tweet.replace('>', '')
        tweet = tweet.replace('?', '')
        tweet = tweet.replace('\(', ' \( ')
        tweet = tweet.replace('\)', ' \) ')
        return tweet

    @staticmethod
    def is_word(word):
        if re.search(r'^\w+$',word) and not re.search(r'^\d+', word):
            return True
        else:
            return False

    @staticmethod
    def remove_noise(tokens, dict_noise):
        tokens2 = []
        for ii in range(0, len(tokens)):
            if (tokens[ii] not in dict_noise):
                tokens2.append(tokens[ii])
        return tokens2

    @staticmethod
    def preprocessing (tweet, dict_noise):
        if (type(tweet) == int) or (type(tweet) == long) or (type(tweet) == float) or (type(tweet) == complex):
            tweet = str(tweet)

        tweet = re.sub(r'\'', "", tweet)
        tweet = Preprocessing.remove_username(tweet)
        tweet = Preprocessing.remove_hashtag(tweet)
        tweet = Preprocessing.remove_url(tweet)

        tweet = tweet.lower()
        tweet = Preprocessing.remove_sign(tweet)

        tokens = tweet.split()

        for i in range(0, len(tokens)):
            tokens[i] = Preprocessing.stemming(tokens[i])
            tokens[i] = Preprocessing.convert_numberToLetter(tokens[i])
            tokens[i] = Preprocessing.convert_symbolToLetter(tokens[i])
            tokens[i] = Preprocessing.remove_multipleLetter(tokens[i])

        result = Preprocessing.remove_noise(tokens, dict_noise)
        return result

class FeatureSelection:

    @staticmethod
    def mutualInformation():
        nb = openpyxl.Workbook()
        ws = nb.active

        bp = xlrd.open_workbook('dataset_preprocessing.xlsx')
        dataPreproc = bp.sheet_by_index(0)
        rowLen = dataPreproc.nrows

        totalToken = 0
        totalLabel = 1
        label = []

        for i in range(0, rowLen):
            data = dataPreproc.cell(i, 0).value
            tokens = data.split()

            if (i == 0):
                label.append(dataPreproc.cell(i, 1).value)

            if (i>0 and (dataPreproc.cell(i, 1).value != dataPreproc.cell(i-1, 1).value)):
                totalLabel += 1
                label.append(dataPreproc.cell(i, 1).value)

            for j in range(0, len(tokens)):
                totalToken += 1

        tokenOnLabel = []

        for i in range(0, len(label)):
            for j in range(0, rowLen):
                if(dataPreproc.cell(j, 1).value == label[i]):
                    data = dataPreproc.cell(j, 0).value
                    tokens = data.split()
                    for k in range(0, len(tokens)):
                        tokenOnLabel.append([])
                        tokenOnLabel[i].append(tokens[k])

        ws.append(['TOKEN', 'LABEL', 'N11', 'N01', 'N10', 'N00', 'I(U;C)'])
        for i in range(0, len(label)):
            data = set(tokenOnLabel[i])
            data = list(data)

            for j in range(0, len(data)):

                n10 = 0
                n00 = 0
                for k in range(0, len(label)):
                    if (label[i] != label[k]):
                        n10 += tokenOnLabel[k].count(data[j])

                # for k in range(0, len(label)):
                #     if (label[i] != label[k]):
                #
                #             n00 += tokenOnLabel[k].count(not (data[j]))


                n11 = tokenOnLabel[i].count(data[j])
                n01 = len(tokenOnLabel[i]) - n11
                n00 = totalToken - n11 - n10 - n01

                a = ((n11/totalToken) * math.log((totalToken*n11 / ((n11+n10)*(n11+n01)) ), 2))
                b = ((n01/totalToken) * math.log((totalToken*n01 / ((n01+n00)*(n01+n11)) ), 2))
                d = ((n00/totalToken) * math.log((totalToken*n00 / ((n00+n01)*(n00+n10)) ), 2))

                if((n10) == 0):
                    c = 0
                    iuc = a + b - c + d
                    # iuc = math.log( ((n11 * totalToken) / (n11 + n10) * (n11 + n01)), 2 )
                else:
                    c = ((n10/totalToken) * math.log((totalToken*n10) / ((n10+n11)*(n10+n00)), 2))
                    iuc = a + b - c + d
                # iuc = math.log( ((n11 * totalToken) / ((n11 + n10) * (n11 + n01))), 2 )

                # print n11, n10, n01, n00, totalToken, iuc
                # print a, b, c, d

                ws.append([data[j], label[i], n11, n01, n10, n00, iuc])
        nb.save("dataset_MIvalue.xlsx")

    @staticmethod
    def elimination():
        nb = openpyxl.Workbook()
        ws = nb.active

        wf = xlrd.open_workbook('dataset_MIvalue.xlsx')
        dataFeature = wf.sheet_by_index(0)
        rowLen = dataFeature.nrows-1

        label = []
        term = []
        n = []
        avg = []

        for i in range(0, rowLen):
            label.append(dataFeature.cell(i+1, 1).value)
            term.append(dataFeature.cell(i+1, 0).value)
        labelSet = set(label)
        labelSet = list(labelSet)
        termSet = set(term)
        termSet = list(termSet)

        #calc average avg per class
        for i in range(0, len(labelSet)):
            n.append(0)
            totalMIvalue = 0
            for j in range(0, rowLen):
                if (dataFeature.cell(j+1, 1).value == labelSet[i]):
                    n[i] += 1
                    totalMIvalue += dataFeature.cell(j+1, 6).value
            avg.append(totalMIvalue / n[i])
            print labelSet[i], totalMIvalue, n[i], avg[i]

        ws.append(['TOKEN', 'LABEL', 'N11', 'N01', 'N10', 'N00', 'I(U;C)', 'AVG'])
        k = 2
        for i in range(0, len(labelSet)):
            for j in range(0, rowLen):
                if ((dataFeature.cell(j+1, 1).value == labelSet[i]) and (dataFeature.cell(j+1, 6).value >= avg[i])):
                    ws.cell(row=k, column=1).value = dataFeature.cell(j+1, 0).value
                    ws.cell(row=k, column=2).value = dataFeature.cell(j+1, 1).value
                    ws.cell(row=k, column=3).value = dataFeature.cell(j+1, 2).value
                    ws.cell(row=k, column=4).value = dataFeature.cell(j+1, 3).value
                    ws.cell(row=k, column=5).value = dataFeature.cell(j+1, 4).value
                    ws.cell(row=k, column=6).value = dataFeature.cell(j+1, 5).value
                    ws.cell(row=k, column=7).value = dataFeature.cell(j+1, 6).value
                    ws.cell(row=k, column=8).value = avg[i]
                    k += 1
        nb.save('dataset_selectedFeature.xlsx')

class Classification:
    @staticmethod
    def naiveBayes():
        nb = openpyxl.Workbook()
        ws = nb.active

        wt = xlrd.open_workbook('DATA TRAIN.xlsx')
        dataTrain = wt.sheet_by_index(0)
        rowLenTrain = dataTrain.nrows

        wf = xlrd.open_workbook('dataset_selectedFeature.xlsx')
        dataFeature = wf.sheet_by_index(0)
        rowLen = dataFeature.nrows-1

        label = []
        term = []
        c = []

        for i in range(0, rowLenTrain):
            label.append(dataTrain.cell(i, 1).value)

        # get label and term in vocab
        for i in range(0, rowLen):
            # label.append(dataFeature.cell(i+1, 1).value)
            term.append(dataFeature.cell(i+1, 0).value)

        labelSet = set(label)
        labelSet = list(labelSet)
        termSet = set(term)
        termSet = list(termSet)

        # get pc(probability of class)
        # set label and pc in xls
        ws.append(['Data Probabilitas'])
        ws.append(['Prob Class'])
        for i in range(0, len(labelSet)):
            ws.cell(row=1, column=i+2).value = labelSet[i]
            c.append([])
            c[i].append(labelSet[i])
            c[i].append(label.count(labelSet[i]))
            pc = c[i][1] / rowLenTrain
            # pc = c[i][1] / len(termSet)
            ws.cell(row=2, column=i+2).value = pc

        # set term probabilty of classes
        for i in range(0, len(termSet)):
            # print 'i', i, termSet[i]
            ws.append([termSet[i]])
            for j in range(0, len(labelSet)):
                # print '  j', j, ws.cell(row=1, column=j+2).value
                endLoop = False
                k = 0
                while(endLoop == False):

                    if (k == dataFeature.nrows-1):
                        n11 = 0
                        endLoop = True
                        break
                    elif ((termSet[i] == dataFeature.cell(k+1, 0).value) and (ws.cell(row=1, column=j+2).value == dataFeature.cell(k+1, 1).value)):
                        n11 = dataFeature.cell(k+1, 2).value
                        endLoop = True
                        break
                    # print '    k',k, dataFeature.cell(k+1, 0).value, dataFeature.cell(k+1, 1)
                    k += 1

                for l in range(0, len(c)):
                    # print '    l', l, c[l][0]
                    if(c[l][0] == ws.cell(row=1, column=j+2).value):
                        termInClass = c[l][1]
                p = (n11 + 1) / (termInClass + len(termSet))
                ws.cell(row=3+i, column=j+2).value = p

        nb.save("data_classification_MI.xlsx")

class Training:
    @staticmethod
    def training():
        dict_noise = DictionaryHandler.loadDict_noise()

        wb = xlrd.open_workbook('DATA TRAIN.xlsx')
        dataTrain = wb.sheet_by_index(0)
        rowLen = dataTrain.nrows

        nb = openpyxl.Workbook();
        ns = nb.active

        for i in range(0, rowLen):
            data_i = dataTrain.cell(i,0).value
            class_i = dataTrain.cell(i, 1).value
            prep = Preprocessing.preprocessing(data_i, dict_noise)

            if prep:
                ns.append([' '.join(prep), class_i])

        nb.save("dataset_preprocessing.xlsx")

        FeatureSelection.mutualInformation()
        FeatureSelection.elimination()
        Classification.naiveBayes()

class Testing:
    @staticmethod
    def oneTweet(tweet):
        # load dictionary noise
        dict_noise = DictionaryHandler.loadDict_noise()

        # load xls classification
        cw = xlrd.open_workbook("data_classification_nonMI.xlsx")
        dataClassification = cw.sheet_by_index(0)

        label = []
        # preprocessing for data input
        token = Preprocessing.preprocessing(tweet, dict_noise)

        # get label from xls
        for i in range(1, dataClassification.ncols):
            label.append(dataClassification.cell(0, i).value)

        if (token):
            p = []
            # get probability
            i = 0
            idx = 0
            while(i < len(token)):
                p.append([])
                j = 2
                while(j < dataClassification.nrows):
                    if(token[i] == dataClassification.cell(j, 0).value):
                        for k in range(0, len(label)):
                            pd = dataClassification.cell(j, k+1).value
                            p[idx].append(pd)
                        break
                    j += 1
                if (j == dataClassification.nrows):
                    del p[-1]
                    i += 1
                else:
                    i += 1
                    idx += 1

            # probability calc
            pFinal = []
            for i in range(0, len(label)):
                pc = dataClassification.cell(1, i+1).value
                valP = 1
                for j in range(0, len(p)):
                    valP *= p[j][i]

                if(len(p) == 1):
                    if(valP != 1):
                        value = valP
                    else:
                        value = 0
                else:
                    if(valP != 1):
                        value = valP*pc
                    else:
                        value = 0

                pFinal.append(value)

            maks = max(pFinal)

            if(maks != 0):
                for i in range(0, len(pFinal)):
                    if(pFinal[i] == maks):
                        idxMax = i
                decision = dataClassification.cell(0, idxMax+1).value
            else:
                decision = 'ERROR'
        else:
            decision = 'pribadi'

        return decision


    @staticmethod
    def oneFile(fileName):
        nb = openpyxl.Workbook();
        ns = nb.active

        tw = xlrd.open_workbook(fileName)
        dataTest = tw.sheet_by_index(0)
        dataLen = dataTest.nrows-1

        decisionTrue = 0
        decisionError = 0
        allCategory = []
        allDec = []

        for i in range(1, dataLen+1):
            print i
            tweet = dataTest.cell(i, 0).value
            decision_actually = dataTest.cell(i, 1).value
            decision_test = Testing.oneTweet(tweet)

            # print decision_actually, decision_test


            # print decision_actually, '----------', decision_test
            ns.append([tweet, decision_actually, decision_test])

            if(decision_test == 'ERROR'):
                decisionError += 1
            elif(decision_actually == decision_test):
                decisionTrue += 1

            allCategory.append(decision_actually)
            allDec.append(decision_test)

        cm = ConfusionMatrix(allCategory, allDec)
        print cm

        accuracy = (decisionTrue / (dataLen - decisionError)) * 100
        ns.append([' ', 'DECISION TRUE', 'DECISION FALSE', 'DECISION ERROR', 'ACCURACY'])
        ns.append([' ', decisionTrue, (dataLen - decisionError - decisionTrue), decisionError, accuracy])
        nb.save("RESULT_nonMI_train-train.xlsx")

# Training.training()

# FeatureSelection.mutualInformation()
# FeatureSelection.elimination()
# Classification.naiveBayes()

Testing.oneFile('DATA TRAIN.xlsx')

# tweet = '@Izs_Turuy Tujuannya bukan hanya sebatas presiden, sebetulnya citra itu sangat penting, bisa saja dia lengser jika tak memiliki citra'
# print Testing.oneTweet(tweet)

# dict_noise = DictionaryHandler.loadDict_noise()
# print Preprocessing.preprocessing(tweet, dict_noise)
